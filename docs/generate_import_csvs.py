"""Generate import-ready CSVs from docs/semenyih_db.xlsx using docs/templates/* schemas."""
import openpyxl
import csv
import os
import re
from collections import defaultdict
from datetime import datetime, date, time

SRC = r'C:\Users\khsra\OneDrive\Documents\GitHub\advaspire-system\docs\semenyih_db.xlsx'
OUT_DIR = r'C:\Users\khsra\OneDrive\Documents\GitHub\advaspire-system\docs\import-csv'
os.makedirs(OUT_DIR, exist_ok=True)

BRANCH_NAME = "Advaspire Robotics & Coding Academy Semenyih"
IMPORTER_EMAIL = "advaspire@gmail.com"

wb = openpyxl.load_workbook(SRC, data_only=True)


# ---------- helpers ----------
def fmt_date(v):
    if v is None or v == "":
        return ""
    if isinstance(v, datetime):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, date):
        return v.strftime("%Y-%m-%d")
    s = str(v).strip()
    # already YYYY-MM-DD?
    if re.match(r"^\d{4}-\d{2}-\d{2}$", s):
        return s
    return ""


def fmt_time(v):
    if v is None or v == "":
        return ""
    if isinstance(v, time):
        return v.strftime("%H:%M")
    if isinstance(v, datetime):
        return v.strftime("%H:%M")
    s = str(v).strip()
    m = re.match(r"^(\d{1,2}):(\d{2})", s)
    if m:
        return f"{int(m.group(1)):02d}:{m.group(2)}"
    return ""


def clean_phone(s):
    if not s:
        return ""
    return re.sub(r"\s+", "", str(s).strip())


def safe_str(v):
    if v is None:
        return ""
    return str(v).strip()


def num(v, default=0):
    if isinstance(v, (int, float)):
        return int(v)
    try:
        return int(float(v))
    except (TypeError, ValueError):
        return default


# ---------- load lookups ----------
customers = {}
ws = wb["customer_database"]
for row in ws.iter_rows(min_row=2, values_only=True):
    cid, name, contact, email, sex, a1, a2, a3, city, postcode, state, country = row[:12]
    if not cid:
        continue
    addr_parts = [safe_str(a1), safe_str(a2), safe_str(a3)]
    address = ", ".join([p for p in addr_parts if p])
    customers[cid] = {
        "name": safe_str(name),
        "contact": clean_phone(contact),
        "email": safe_str(email),
        "address": address,
        "city": safe_str(city),
    }

parent_to_customer = {}
ws = wb["customer_parent_map"]
for row in ws.iter_rows(min_row=2, values_only=True):
    pid, cid, ic = row[:3]
    if pid and cid:
        parent_to_customer[pid] = cid

ticket_balance = {}
ws = wb["ticket_balance"]
for row in ws.iter_rows(min_row=2, values_only=True):
    sid, sname, active, r_bal, r_used, r_remain, c_bal, c_used, c_remain = row[:9]
    if not sid:
        continue
    ticket_balance[sid] = {
        "active": (str(active).strip().lower() == "active") if active else False,
        "r_bal": num(r_bal),
        "r_remain": num(r_remain),
        "c_bal": num(c_bal),
        "c_remain": num(c_remain),
    }

students = {}
ws = wb["student_database"]
for row in ws.iter_rows(min_row=2, values_only=True):
    enroll_date, sid, pid, sname, idcard, ic, birth, born_year, school, ethn, state, sex, joint = row[:13]
    if not sid:
        continue
    students[sid] = {
        "name": safe_str(sname),
        "parent_id": pid,
        "dob": fmt_date(birth) if birth else "",
        "gender": safe_str(sex).lower() if sex else "",
        "school": safe_str(school),
        "enroll_date": fmt_date(enroll_date) if enroll_date else "",
    }


def fallback_email(parent_id, parent_name):
    base = re.sub(r"[^a-z0-9]+", ".", parent_name.lower()).strip(".")
    if not base:
        base = f"parent.{(parent_id or 'unknown').lower()}"
    return f"{base}.{(parent_id or 'unknown').lower()}@imported.local"


# ---------- 1. students.csv ----------
header_students = [
    "student_id", "student_name", "date_of_birth", "gender", "school_name",
    "parent_name", "parent_email", "parent_phone", "parent_address", "parent_city",
    "branch_name", "program_name", "package_type", "package_duration",
    "schedule_day", "schedule_time", "enrollment_status", "sessions_remaining",
    "share_with_sibling",
]

students_rows = []
missing_parent = 0
# Track parent_email per parent_id so siblings end up with the same value
# (the import dedups parents by lowercased email).
parent_email_lookup = {}
for sid, s in students.items():
    pid = s["parent_id"]
    cust = customers.get(parent_to_customer.get(pid)) if pid else None
    if not cust:
        parent_name = "Unknown Parent"
        parent_email = parent_email_lookup.get(pid) or fallback_email(pid or sid, parent_name)
        parent_phone = ""
        parent_addr = ""
        parent_city = ""
        missing_parent += 1
    else:
        parent_name = cust["name"] or "Unknown Parent"
        parent_email = cust["email"] or parent_email_lookup.get(pid) or fallback_email(pid, parent_name)
        parent_phone = cust["contact"]
        parent_addr = cust["address"]
        parent_city = cust["city"]
    if pid:
        parent_email_lookup[pid] = parent_email

    tb = ticket_balance.get(sid)
    enrolls = []
    if tb:
        status = "active" if tb["active"] else "inactive"
        if tb["r_bal"] != 0 or tb["r_remain"] != 0:
            enrolls.append(("Lego EV3 Robotics", status, max(tb["r_remain"], 0)))
        if tb["c_bal"] != 0 or tb["c_remain"] != 0:
            enrolls.append(("Scratch", status, max(tb["c_remain"], 0)))

    base_row = {
        "student_id": sid,
        "student_name": s["name"],
        "date_of_birth": s["dob"],
        "gender": s["gender"],
        "school_name": s["school"],
        "parent_name": parent_name,
        "parent_email": parent_email,
        "parent_phone": parent_phone,
        "parent_address": parent_addr,
        "parent_city": parent_city,
        "branch_name": BRANCH_NAME,
        "package_type": "",
        "package_duration": "",
        "schedule_day": "",
        "schedule_time": "",
        "share_with_sibling": "",
    }
    if not enrolls:
        r = dict(base_row)
        r.update({"program_name": "", "enrollment_status": "", "sessions_remaining": ""})
        students_rows.append(r)
    else:
        for program, status, sessions in enrolls:
            r = dict(base_row)
            r.update({
                "program_name": program,
                "enrollment_status": status,
                "sessions_remaining": str(sessions),
            })
            students_rows.append(r)

with open(os.path.join(OUT_DIR, "students.csv"), "w", newline="", encoding="utf-8") as f:
    w = csv.DictWriter(f, fieldnames=header_students)
    w.writeheader()
    for r in students_rows:
        w.writerow(r)
print(f"students.csv: {len(students_rows)} rows from {len(students)} students ({missing_parent} missing parent info)")


# ---------- 2. attendance.csv ----------
# Source: attendance_history. Drops rows for students not in our students table
# (they won't resolve at import time anyway).
header_attendance = [
    "student_id", "date", "time", "status", "class_type", "instructor_name",
    "lesson", "mission", "activity", "adcoin", "absence_reason",
]

def normalize_instructor(name):
    if not name:
        return ""
    return safe_str(name).title()


status_map = {"present": "present", "absent": "absent", "late": "late", "excused": "excused"}

attendance_rows = []
ws = wb["attendance_history"]
skipped_no_student = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    no, uid, name, course, klass, dt, day, tm, attendance, activity, adcoins, materials, teacher = row[:13]
    if not uid:
        continue
    if uid not in students:
        skipped_no_student += 1
        continue
    s = status_map.get(safe_str(attendance).lower(), "present")
    ctype = safe_str(klass)
    if ctype.lower() == "physical":
        ctype = "Physical"
    elif ctype.lower() == "online":
        ctype = "Online"
    else:
        ctype = ""
    coins_val = ""
    if isinstance(adcoins, (int, float)):
        coins_val = str(int(adcoins))
    attendance_rows.append({
        "student_id": uid,
        "date": fmt_date(dt),
        "time": fmt_time(tm),
        "status": s,
        "class_type": ctype,
        "instructor_name": normalize_instructor(teacher),
        "lesson": "",
        "mission": "",
        "activity": safe_str(activity),
        "adcoin": coins_val,
        "absence_reason": "",
    })

with open(os.path.join(OUT_DIR, "attendance.csv"), "w", newline="", encoding="utf-8") as f:
    w = csv.DictWriter(f, fieldnames=header_attendance)
    w.writeheader()
    for r in attendance_rows:
        w.writerow(r)
print(f"attendance.csv: {len(attendance_rows)} rows ({skipped_no_student} skipped — student not in DB)")


# ---------- 3. payments.csv ----------
# Source: payment_history. Course → program_name mapping below.
header_payments = [
    "student_id", "amount", "status", "payment_method", "paid_at",
    "program_name", "package_type", "package_duration", "receipt_seq", "notes",
]

course_map_payments = {
    "robotics class": "Lego EV3 Robotics",
    "coding class": "Scratch",
    "scratch coding": "Scratch",
    "bootcamp": "",
    "free trial": "",
    "advaspire shirt": "",
    "advasbot e3 set": "",
    "competition": "",
}

# payment_method guess from reference column
def guess_method(ref):
    r = (safe_str(ref) or "").lower()
    if not r:
        return ""
    if "cash" in r:
        return "cash"
    if "qr" in r or "duitnow" in r or "duit now" in r:
        return "promptpay"
    if "card" in r or "credit" in r:
        return "credit_card"
    if "transfer" in r or "tng" in r or "boost" in r or "bank" in r:
        return "bank_transfer"
    return ""


payment_rows = []
ws = wb["payment_history"]
skipped_payments = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    dt, fmt_d, ticket_id, running_no, cid, sid, sname, pid, ref, amount, ticket, course, month, year = row[:14]
    if not sid:
        continue
    if sid not in students:
        skipped_payments += 1
        continue
    if not isinstance(amount, (int, float)):
        continue
    if amount <= 0:
        # zero / negative entries (adjustments) — skip; the import requires
        # amount > 0 isn't enforced but $0 records aren't useful here.
        continue
    course_lower = safe_str(course).lower()
    program = course_map_payments.get(course_lower, "")
    notes_parts = []
    if ticket_id:
        notes_parts.append(f"old ticket {ticket_id}")
    if ref:
        notes_parts.append(safe_str(ref))
    notes = " | ".join(notes_parts)
    payment_rows.append({
        "student_id": sid,
        "amount": f"{float(amount):.2f}",
        "status": "paid",
        "payment_method": guess_method(ref),
        "paid_at": fmt_date(dt),
        "program_name": program,
        "package_type": "",
        "package_duration": "",
        "receipt_seq": "",
        "notes": notes,
    })

with open(os.path.join(OUT_DIR, "payments.csv"), "w", newline="", encoding="utf-8") as f:
    w = csv.DictWriter(f, fieldnames=header_payments)
    w.writeheader()
    for r in payment_rows:
        w.writerow(r)
print(f"payments.csv: {len(payment_rows)} rows ({skipped_payments} skipped — student not in DB)")


# ---------- 4. examinations.csv ----------
# Source: Cert_Database. Level string → integer.
header_exams = [
    "student_id", "exam_name", "exam_level", "exam_date", "status", "mark",
    "reattempt_count", "examiner_email", "certificate_number", "notes",
]

level_int = {
    "junior": 0,
    "beginner": 1,
    "level 1": 1,
    "intermediate": 2,
    "level 2": 2,
    "advanced": 3,
    "level 3": 3,
}

subject_full_name = {
    "ev3": "EV3 Robotics",
    "scratch": "Scratch",
    "robotic": "Robotics",
    "python": "Python",
    "minecraft": "Minecraft",
}

exam_rows = []
ws = wb["Cert_Database"]
skipped_exams = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    no, start, sid, ic, fn, ln, full_name, remark, program, subject, full_course, level, completion, result, grade, run_no, cert_no, accessor, pdf, printed = row[:20]
    if not sid:
        continue
    if sid not in students:
        skipped_exams += 1
        continue
    lvl_key = safe_str(level).lower()
    lvl_int = level_int.get(lvl_key)
    if lvl_int is None:
        # Unmappable level — skip rather than guess; the import requires a
        # numeric level.
        skipped_exams += 1
        continue
    subj = subject_full_name.get(safe_str(subject).lower(), safe_str(subject))
    exam_name = f"{subj} {safe_str(level)}".strip()
    if not exam_name:
        continue
    status = "pass"  # all rows in this file have a grade — treated as pass
    mark_val = ""
    if isinstance(result, (int, float)):
        # result stored as 0.00–1.00 — convert to integer percentage.
        mark_pct = int(round(float(result) * 100))
        mark_val = str(mark_pct)
    exam_rows.append({
        "student_id": sid,
        "exam_name": exam_name,
        "exam_level": str(lvl_int),
        "exam_date": fmt_date(completion),
        "status": status,
        "mark": mark_val,
        "reattempt_count": "0",
        "examiner_email": "",
        "certificate_number": safe_str(cert_no),
        "notes": f"Program: {safe_str(program)} | Grade: {safe_str(grade)}",
    })

with open(os.path.join(OUT_DIR, "examinations.csv"), "w", newline="", encoding="utf-8") as f:
    w = csv.DictWriter(f, fieldnames=header_exams)
    w.writeheader()
    for r in exam_rows:
        w.writerow(r)
print(f"examinations.csv: {len(exam_rows)} rows ({skipped_exams} skipped)")


# ---------- 5. transactions.csv ----------
# Source: adcoins_history. Earned (Mission/Activity/Competition/Level
# Completion/Others) → sender=importer (user), receiver=student.
# Purchase → sender=student, receiver=importer (user), type=spent.
header_tx = [
    "sender_id", "sender_type", "receiver_id", "receiver_type", "amount",
    "type", "date", "description",
]

tx_rows = []
ws = wb["adcoins_history"]
skipped_tx = 0
zero_amount = 0
for row in ws.iter_rows(min_row=2, values_only=True):
    tx_id, dt, uid, sname, category, amount, description, remark, auth_by, year, month = row[:11]
    if not uid:
        continue
    if uid not in students:
        skipped_tx += 1
        continue
    if not isinstance(amount, (int, float)):
        continue
    amt = int(amount)
    if amt <= 0:
        zero_amount += 1
        continue
    cat = safe_str(category).lower()
    if cat == "purchase":
        sender_id = uid
        sender_type = "student"
        receiver_id = IMPORTER_EMAIL
        receiver_type = "user"
        tx_type = "spent"
    else:
        sender_id = IMPORTER_EMAIL
        sender_type = "user"
        receiver_id = uid
        receiver_type = "student"
        tx_type = "earned"
    desc_parts = [safe_str(category)]
    if description:
        desc_parts.append(safe_str(description))
    if remark:
        desc_parts.append(safe_str(remark))
    if auth_by:
        desc_parts.append(f"by {safe_str(auth_by)}")
    tx_rows.append({
        "sender_id": sender_id,
        "sender_type": sender_type,
        "receiver_id": receiver_id,
        "receiver_type": receiver_type,
        "amount": str(amt),
        "type": tx_type,
        "date": fmt_date(dt),
        "description": " | ".join([p for p in desc_parts if p]),
    })

with open(os.path.join(OUT_DIR, "transactions.csv"), "w", newline="", encoding="utf-8") as f:
    w = csv.DictWriter(f, fieldnames=header_tx)
    w.writeheader()
    for r in tx_rows:
        w.writerow(r)
print(f"transactions.csv: {len(tx_rows)} rows ({skipped_tx} skipped — student not in DB, {zero_amount} zero-amount skipped)")

print("\nAll CSVs written to:", OUT_DIR)
